import xlwt
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render, redirect
from .models import *
from django.urls import reverse
from hrm.models import EMP_POLICE_VERFICATION, Country, Client, New_Emp, salary_details, Monthly_Attendance_Table, \
    EMP_COMPANY_DETAILS, Loan_Table

import csv

from django.http import HttpResponse
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import json as simplejson

from hrm.views import find_emp_name, find_emp_monthly_atten


def file_load_view(request):
    response = HttpResponse(content_type='text/xlsx')
    response['Content-Disposition'] = 'attachement; filename="report.xlsx"'

    writer = csv.writer(response)
    writer.writerow(['Student Name', 'Attendance'])

    students = Country.objects.values('Country_Name').annotate(mark=Sum('ID'))

    # Note: we convert the students query set to a values_list as the writerow expects a list/tuple
    students = students.values_list('Country_Name', 'ID')

    for student in students:
        writer.writerow(student)

    return response


@login_required(login_url='login')
def form_a(request):
    if request.method == "POST":
        client_list = []
        country_list = []
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_client = request.POST.get('client')
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)
        if not data:
            filter_value = "no"
        value={
            'data': data,
            'client': client_list,
            'error_message': data,
            'country_list':country_list,
            'filter_value':filter_value,
            'dummy_value':filter_client,
        }
        return render(request, 'Form_A.html',value )

    else:
        if request.GET.get('client') is None:
            newlist = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in newlist:
                    newlist.append(variable)
            print(newlist)
            form = request.POST
            stu = {
                'client': newlist,
            }
        else:
            newlist = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in newlist:
                    newlist.append(variable)
            filter_value = "yes"
            client = request.GET.get('client')
            result_set = []
            filter_client = CLIENT_DETAILS.objects.filter(CLIENT=client)
            for country in filter_client:
                result_set.append(country.COUNTRY_NAME)

            for i in range(len(result_set)):
                country_list = [result_set[i]]
                for e in result_set:
                    if e not in country_list:
                        country_list.append(e)
            print(newlist)
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=client)
            stu = {
                'data': data,
                'client': newlist,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': client,
            }
    return render(request, 'Form_A.html', stu)


@login_required(login_url='login')
def Filter_Form_A(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        filter_client = request.POST.get('client_name')
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+request.POST.get('client_name')+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+request.POST.get('client_name')+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+request.POST.get('client_name')+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+request.POST.get('client_name')+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+request.POST.get('client_name')+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+request.POST.get('client_name')+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        value = {
            'data': data,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'dummy_value': filter_client,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'Form_A.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_A'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable, "variable")
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2],
                    EMP_COMPANY__STATE=get_value[3])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2],
                    EMP_COMPANY__STATE=get_value[3],
                    EMP_COMPANY__DISTRICT=get_value[4])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2],
                    EMP_COMPANY__STATE=get_value[3],
                    EMP_COMPANY__DISTRICT=get_value[4],
                    EMP_COMPANY__BRANCH=get_value[5])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2],
                    EMP_COMPANY__STATE=get_value[3],
                    EMP_COMPANY__DISTRICT=get_value[4],
                    EMP_COMPANY__BRANCH=get_value[5],
                    EMP_COMPANY__AREA=get_value[6])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[2],
                    EMP_COMPANY__STATE=get_value[3],
                    EMP_COMPANY__DISTRICT=get_value[4],
                    EMP_COMPANY__BRANCH=get_value[5],
                    EMP_COMPANY__AREA=get_value[6],
                    EMP_COMPANY__UNIT=get_value[7])
            value = {
                'data': data,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': filter_client,
                'form_a_filter': form_a_filter,
            }
            return render(request, 'Form_A.html', value)

@login_required(login_url='login')
def Form_D(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        emp_count=0
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    Designation = emp_details.EMP_COMPANY.DESIGNATION
                    Acc_No = emp_details.EMP_BANK.ACCOUNT_NO
                    IFSC_Code = emp_details.EMP_BANK.IFSC_CODE
                    Net_Payment = data1.NET_PAY

                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'Form_D.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"

                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        Designation = emp_details.EMP_COMPANY.DESIGNATION
                        Acc_No = emp_details.EMP_BANK.ACCOUNT_NO
                        IFSC_Code = emp_details.EMP_BANK.IFSC_CODE
                        Net_Payment = data1.NET_PAY

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'Form_D.html', value)

@login_required(login_url='login')
def Filter_Form_D(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    Designation = emp_details.EMP_COMPANY.DESIGNATION
                    Acc_No = emp_details.EMP_BANK.ACCOUNT_NO
                    IFSC_Code = emp_details.EMP_BANK.IFSC_CODE
                    Net_Payment = data1.NET_PAY

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'Form_D.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_D'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        Designation = emp_details.EMP_COMPANY.DESIGNATION
                        Acc_No = emp_details.EMP_BANK.ACCOUNT_NO
                        IFSC_Code = emp_details.EMP_BANK.IFSC_CODE
                        Net_Payment = data1.NET_PAY

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'Form_D.html', value)
@login_required(login_url='login')
def Form_23(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    EMP_CODE = data1.EMP_CODE
                    EMP_NAME = emp_details.NEW_EMP.EMP_NAME
                    FATHER_NAME = emp_details.EMP_PER.FATHER_NAME
                    SEX = emp_details.NEW_EMP.GENDER
                    DESIGNATION = emp_details.EMP_COMPANY.DESIGNATION
                    DATE = request.POST.get("date_month").split("-")
                    OVERTIME = data1.OVERTIME_HRS
                    NORMAL_RATE_WAGES = data1.OVERTIME_AMOUNT

                    EARNINGS = data1.OVERTIME_AMOUNT

                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'emp_overtime_form.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        EMP_CODE = data1.EMP_CODE
                        EMP_NAME = emp_details.NEW_EMP.EMP_NAME
                        FATHER_NAME = emp_details.EMP_PER.FATHER_NAME
                        SEX = emp_details.NEW_EMP.GENDER
                        DESIGNATION = emp_details.EMP_COMPANY.DESIGNATION
                        DATE = get_value[1].split("-")
                        OVERTIME = data1.OVERTIME_HRS
                        NORMAL_RATE_WAGES = data1.OVERTIME_AMOUNT

                        EARNINGS = data1.OVERTIME_AMOUNT

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'emp_overtime_form.html', value)

@login_required(login_url='login')
def Filter_Form_23(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    EMP_CODE = data1.EMP_CODE
                    EMP_NAME = emp_details.NEW_EMP.EMP_NAME
                    FATHER_NAME = emp_details.EMP_PER.FATHER_NAME
                    SEX = emp_details.NEW_EMP.GENDER
                    DESIGNATION = emp_details.EMP_COMPANY.DESIGNATION
                    DATE = request.POST.get("date_month").split("-")
                    OVERTIME = data1.OVERTIME_HRS
                    NORMAL_RATE_WAGES = data1.OVERTIME_AMOUNT

                    EARNINGS = data1.OVERTIME_AMOUNT

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'emp_overtime_form.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_B'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        EMP_CODE = data1.EMP_CODE
                        EMP_NAME = emp_details.NEW_EMP.EMP_NAME
                        FATHER_NAME = emp_details.EMP_PER.FATHER_NAME
                        SEX = emp_details.NEW_EMP.GENDER
                        DESIGNATION = emp_details.EMP_COMPANY.DESIGNATION
                        DATE = get_value[2].split("-")
                        OVERTIME = data1.OVERTIME_HRS
                        NORMAL_RATE_WAGES = data1.OVERTIME_AMOUNT

                        EARNINGS = data1.OVERTIME_AMOUNT

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'emp_overtime_form.html', value)

@login_required(login_url='login')
def form_pf(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    emp_code = data1.EMP_CODE
                    total_contributions = float(data1.PENSION_AMOUNT) + float(data1.PROVIDENT_FUND)
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    uan_no = emp_details.EMP_PER.UAN_NO
                    BASIC = data1.BASIC
                    pension_amount = data1.PENSION_AMOUNT
                    PF = data1.PROVIDENT_FUND
                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'form_pf.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        emp_code = data1.EMP_CODE
                        total_contributions = float(data1.PENSION_AMOUNT) + float(data1.PROVIDENT_FUND)
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        uan_no = emp_details.EMP_PER.UAN_NO
                        BASIC = data1.BASIC
                        pension_amount = data1.PENSION_AMOUNT
                        PF = data1.PROVIDENT_FUND

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'form_pf.html', value)


@login_required(login_url='login')
def form_pt(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    BASIC = data1.BASIC
                    pt_amount = data1.PROFESSIONAL_TAX
                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'form_pt.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        BASIC = data1.BASIC
                        pt_amount = data1.PROFESSIONAL_TAX

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'form_pt.html', value)


@login_required(login_url='login')
def form_lwf(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    BASIC = data1.BASIC
                    lwf_amount = data1.LABOUR_WELFARE_FUND
                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'form_lwf.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        BASIC = data1.BASIC
                        lwf_amount = data1.LABOUR_WELFARE_FUND

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'form_lwf.html', value)

@login_required(login_url='login')
def form_esi(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    uan_no = emp_details.EMP_PER.UAN_NO
                    month_earn = data1.TOTAL_EARN
                    ESI_NO = emp_details.EMP_PER.ESI_NO
                    ESI = data1.ESIC
                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'form_esi.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        uan_no = emp_details.EMP_PER.UAN_NO
                        month_earn = data1.TOTAL_EARN
                        ESI_NO = emp_details.EMP_PER.ESI_NO
                        ESI = data1.ESIC

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'form_esi.html', value)

@login_required(login_url='login')
def Filter_Form_ESI(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    uan_no = emp_details.EMP_PER.UAN_NO
                    month_earn = data1.TOTAL_EARN
                    ESI_NO = emp_details.EMP_PER.ESI_NO
                    ESI = data1.ESIC

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'form_esi.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_ESI'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        uan_no = emp_details.EMP_PER.UAN_NO
                        month_earn = data1.TOTAL_EARN
                        ESI_NO = emp_details.EMP_PER.ESI_NO
                        ESI = data1.ESIC

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'form_esi.html', value)

@login_required(login_url='login')
def Filter_Form_PF(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    emp_code = data1.EMP_CODE
                    total_contributions = float(data1.PENSION_AMOUNT) + float(data1.PROVIDENT_FUND)
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    uan_no = emp_details.EMP_PER.UAN_NO
                    BASIC = data1.BASIC
                    pension_amount = data1.PENSION_AMOUNT
                    PF = data1.PROVIDENT_FUND

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'form_pf.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_PF'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        emp_code = data1.EMP_CODE
                        total_contributions = float(data1.PENSION_AMOUNT)+float(data1.PROVIDENT_FUND)
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        uan_no = emp_details.EMP_PER.UAN_NO
                        BASIC = data1.BASIC
                        pension_amount = data1.PENSION_AMOUNT
                        PF = data1.PROVIDENT_FUND

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'form_pf.html', value)


@login_required(login_url='login')
def Filter_Form_PT(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    BASIC = data1.BASIC
                    pt_amount = data1.PROFESSIONAL_TAX

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'form_pt.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_PT'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        BASIC = data1.BASIC
                        pt_amount = data1.PROFESSIONAL_TAX

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'form_pt.html', value)


@login_required(login_url='login')
def Filter_Form_LWF(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    emp_code = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    design = emp_details.EMP_COMPANY.DESIGNATION
                    BASIC = data1.BASIC
                    lwf_amount = data1.LABOUR_WELFARE_FUND

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'form_lwf.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_LWF'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        emp_code = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        design = emp_details.EMP_COMPANY.DESIGNATION
                        BASIC = data1.BASIC
                        lwf_amount = data1.LABOUR_WELFARE_FUND

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'form_lwf.html', value)

@login_required(login_url='login')
def Form_B(request):
    if request.method == "POST":
        client_list = []
        report_array = []
        country_list=[]
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable)
            if variable not in client_list:
                client_list.append(variable)
        filter_client=request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        filter_value="no"
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)

        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            print(EMP_CODE)
            employer_details=EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,SALARY_UPDATE_DATE=request.POST.get('date_month'))
            for data1 in employer_details:
                filter_value = "yes"
                print(data1.EMP_CODE)
                class details_class:
                    pay_multi = int(data1.DEARANCE_ALLOWANCES) * int(data1.OVERTIME_HRS)
                    Si_No = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    Rate_of_Wages = data1.FIXED_SALARY
                    No_of_Days_worked = data1.DAYS_PRESENT
                    Overtime_hours_worked = data1.OVERTIME_HRS
                    BASIC = data1.BASIC
                    Special_Basic = 0
                    DA = data1.DEARANCE_ALLOWANCES
                    Payments_overtime = pay_multi
                    HRA = data1.HOUSE_RENT_ALLOWANCES
                    Earn_Others = data1.OTHER_ALLOWANCES
                    Earn_Total = data1.TOTAL_EARN
                    PF = data1.PROVIDENT_FUND
                    ESIC = data1.ESIC
                    Society = 0
                    Income_Tax = data1.INCOME_TAX
                    Insurance = 0
                    Ded_Others = data1.OTHER_DEDUCTION
                    Recoveries = 0
                    Ded_Total = data1.TOTAL_DEDUCATION
                    Net_Payment = data1.NET_PAY
                    Welfare_Fund = 0
                    Transaction_Id = 0
                    Date_of_Payment = 0
                    Remarks = 0

                report_array.append(details_class)

        dummy_value=request.POST.get('client')+"~"+request.POST.get('date_month')
        client_name=request.POST.get('client')
        date_month=request.POST.get('date_month')
        #wage_de = Form_B_Obj(client, unit, month, year)
        value={
            'data': report_array,
            'client': client_list,
            'country_list':country_list,
            'filter_value': filter_value,
            'dummy_value': dummy_value,
            'client_name': client_name,
            'date_month': date_month,
        }

        return render(request, 'Form_B.html', value)

    else:
        if request.GET.get('client') is None:
            client_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            value = {
                'client': client_list,
            }
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            report_array = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in client_list:
                    client_list.append(variable)
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=get_value[0])
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            form = request.POST
            filter_value = "no"
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=get_value[0])

            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                print(EMP_CODE)
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[1])
                for data1 in employer_details:
                    filter_value = "yes"
                    print(data1.FIXED_SALARY,"dd")

                    class details_class:
                        pay_multi = int(data1.DEARANCE_ALLOWANCES) * int(data1.OVERTIME_HRS)
                        Si_No = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        print(Name,"d")
                        Rate_of_Wages = data1.FIXED_SALARY
                        No_of_Days_worked = data1.DAYS_PRESENT
                        Overtime_hours_worked = data1.OVERTIME_HRS
                        BASIC = data1.BASIC
                        Special_Basic = 0
                        DA = data1.DEARANCE_ALLOWANCES
                        Payments_overtime = pay_multi
                        HRA = data1.HOUSE_RENT_ALLOWANCES
                        Earn_Others = data1.OTHER_ALLOWANCES
                        Earn_Total = data1.TOTAL_EARN
                        PF = data1.PROVIDENT_FUND
                        ESIC = data1.ESIC
                        Society = 0
                        Income_Tax = data1.INCOME_TAX
                        Insurance = 0
                        Ded_Others = data1.OTHER_DEDUCTION
                        Recoveries = 0
                        Ded_Total = data1.TOTAL_DEDUCATION
                        Net_Payment = data1.NET_PAY
                        Welfare_Fund = 0
                        Transaction_Id = 0
                        Date_of_Payment = 0
                        Remarks = 0

                    report_array.append(details_class)

            dummy_value = get_value[0]+"~"+get_value[1]
            client_name = get_value[0]
            date_month = get_value[1]
            # wage_de = Form_B_Obj(client, unit, month, year)
            value = {
                'data': report_array,
                'client': client_list,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': dummy_value,
                'client_name': client_name,
                'date_month': date_month,
            }
        return render(request, 'Form_B.html', value)

@login_required(login_url='login')
def Filter_Form_B(request):
    if request.method == "POST":
        client_list=[]
        country_list = []
        report_array=[]
        filter_client = request.POST.get('client_name')
        date_month=request.POST.get('date_month')
        print(filter_client[0],"dyjhgdas")
        client = CLIENT_DETAILS.objects.all()
        form_a_filter=""
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_value = "yes"
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        if request.POST.get('submit')=="Country":
            form_a_filter="Country"+"~"+filter_client+"~"+date_month+"~"+request.POST.get('country_name')
            country_name = request.POST.get('country_name')
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,EMP_COMPANY__COUNTRY=country_name)
        elif request.POST.get('submit')=="State":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            form_a_filter="State"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name)
        elif request.POST.get('submit')=="District":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            form_a_filter="District"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name)
        elif request.POST.get('submit')=="Branch":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name)
        elif request.POST.get('submit')=="Area":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            form_a_filter="Branch"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name)
        elif request.POST.get('submit')=="Unit":
            country_name = request.POST.get('country_name')
            state_name = request.POST.get('state_name')
            district_name = request.POST.get('district_name')
            branch_name = request.POST.get('branch_name')
            area_name = request.POST.get('area_name')
            unit = request.POST.get('unit_name')
            form_a_filter="Unit"+"~"+filter_client+"~"+date_month+"~"+country_name+"~"+state_name+"~"+district_name+"~"+branch_name+"~"+area_name+"~"+unit

            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client,
                                                                                   EMP_COMPANY__COUNTRY=country_name,
                                                                                   EMP_COMPANY__STATE=state_name,
                                                                                   EMP_COMPANY__DISTRICT=district_name,
                                                                                   EMP_COMPANY__BRANCH=branch_name,
                                                                                   EMP_COMPANY__AREA=area_name,
                                                                                   EMP_COMPANY__UNIT=unit)
        for emp_details in data:
            EMP_CODE = emp_details.NEW_EMP.EMP_CODE
            employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                   SALARY_UPDATE_DATE=date_month)
            for data1 in employer_details:
                class details_class:
                    pay_multi = int(data1.DEARANCE_ALLOWANCES) * int(data1.OVERTIME_HRS)
                    Si_No = data1.EMP_CODE
                    Name = emp_details.NEW_EMP.EMP_NAME
                    Rate_of_Wages = data1.FIXED_SALARY
                    No_of_Days_worked = data1.DAYS_PRESENT
                    Overtime_hours_worked = data1.OVERTIME_HRS
                    BASIC = data1.BASIC
                    Special_Basic = 0
                    DA = data1.DEARANCE_ALLOWANCES
                    Payments_overtime = pay_multi
                    HRA = data1.HOUSE_RENT_ALLOWANCES
                    Earn_Others = data1.OTHER_ALLOWANCES
                    Earn_Total = data1.TOTAL_EARN
                    PF = data1.PROVIDENT_FUND
                    ESIC = data1.ESIC
                    Society = 0
                    Income_Tax = data1.INCOME_TAX
                    Insurance = 0
                    Ded_Others = data1.OTHER_DEDUCTION
                    Recoveries = 0
                    Ded_Total = data1.TOTAL_DEDUCATION
                    Net_Payment = data1.NET_PAY
                    Welfare_Fund = 0
                    Transaction_Id = 0
                    Date_of_Payment = 0
                    Remarks = 0

                report_array.append(details_class)
        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'client_name': filter_client,
            'date_month': date_month,
            'form_a_filter':form_a_filter,
        }
        return render(request, 'Form_B.html', value)
    else:
        if request.GET.get('client') is None:
            return redirect(reverse('Form_B'))
        else:
            get_value = request.GET.get('client').split('~')
            client_list = []
            country_list = []
            report_array=[]
            filter_client = get_value[1]
            client = CLIENT_DETAILS.objects.all()
            form_a_filter = ""
            for i in range(len(client)):
                variable = client[i].CLIENT
                if variable not in client_list:
                    client_list.append(variable)
            filter_value = "yes"
            result_set = []
            filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
            for client_details in filter_client_list:
                result_set.append(client_details)

            for i in range(len(result_set)):
                country = result_set[i].COUNTRY_NAME
                if country not in country_list:
                    country_list.append(country)
            if get_value[0] == "Country":
                form_a_filter = "Country" + "~" + get_value[1] + "~" + get_value[2]+ "~" + get_value[3]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3])
            elif get_value[0] == "State":
                form_a_filter = "State" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4])
            elif get_value[0] == "District":
                form_a_filter = "District" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5])
            elif get_value[0] == "Branch":
                form_a_filter = "Branch" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6])
            elif get_value[0] == "Area":
                form_a_filter = "Area" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7])
            elif get_value[0] == "Unit":
                form_a_filter = "Unit" + "~" + get_value[1] + "~" + get_value[2]+"~"+get_value[3]+"~"+get_value[4]+"~"+get_value[5]+"~"+get_value[6]+"~"+get_value[7]+"~"+get_value[8]
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'NEW_EMP').filter(
                    EMP_COMPANY__CLIENT=filter_client,
                    EMP_COMPANY__COUNTRY=get_value[3],
                    EMP_COMPANY__STATE=get_value[4],
                    EMP_COMPANY__DISTRICT=get_value[5],
                    EMP_COMPANY__BRANCH=get_value[6],
                    EMP_COMPANY__AREA=get_value[7],
                    EMP_COMPANY__UNIT=get_value[8])
            for emp_details in data:
                EMP_CODE = emp_details.NEW_EMP.EMP_CODE
                employer_details = EMP_SALARY_MAINTAINS.objects.filter(EMP_CODE=EMP_CODE,
                                                                       SALARY_UPDATE_DATE=get_value[2])
                for data1 in employer_details:
                    class details_class:
                        pay_multi = int(data1.DEARANCE_ALLOWANCES) * int(data1.OVERTIME_HRS)
                        Si_No = data1.EMP_CODE
                        Name = emp_details.NEW_EMP.EMP_NAME
                        Rate_of_Wages = data1.FIXED_SALARY
                        No_of_Days_worked = data1.DAYS_PRESENT
                        Overtime_hours_worked = data1.OVERTIME_HRS
                        BASIC = data1.BASIC
                        Special_Basic = 0
                        DA = data1.DEARANCE_ALLOWANCES
                        Payments_overtime = pay_multi
                        HRA = data1.HOUSE_RENT_ALLOWANCES
                        Earn_Others = data1.OTHER_ALLOWANCES
                        Earn_Total = data1.TOTAL_EARN
                        PF = data1.PROVIDENT_FUND
                        ESIC = data1.ESIC
                        Society = 0
                        Income_Tax = data1.INCOME_TAX
                        Insurance = 0
                        Ded_Others = data1.OTHER_DEDUCTION
                        Recoveries = 0
                        Ded_Total = data1.TOTAL_DEDUCATION
                        Net_Payment = data1.NET_PAY
                        Welfare_Fund = 0
                        Transaction_Id = 0
                        Date_of_Payment = 0
                        Remarks = 0

                    report_array.append(details_class)
            value = {
                'data': report_array,
                'client': client_list,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'client_name': filter_client,
                'date_month': get_value[2],
                'form_a_filter': form_a_filter,
            }
            return render(request, 'Form_B.html', value)


def Form_B_Obj(client, unit, month, year):
    report_array = []

    class details_class:
        Si_No = 0
        Name = 0
        BASIC = 0
        Rate_of_Wages = 0
        No_of_Days_worked = 0
        Overtime_hours_worked = 0
        Basic = 0
        Special_Basic = 0
        DA = 0
        Payments_overtime = 0
        HRA = 0
        Earn_Others = 0
        Earn_Total = 0
        PF = 0
        ESIC = 0
        Society = 0
        Income_Tax = 0
        Insurance = 0
        Ded_Others = 0
        Recoveries = 0
        Ded_Total = 0
        Net_Payment = 0
        Welfare_Fund = 0
        Transaction_Id = 0
        Date_of_Payment = 0
        Remarks = 0

    try:

        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP', 'EMP_SAL').filter(EMP_COMPANY__CLIENT=client)
        for emp_details in data:
            class details_class:
                sal = find_emp_code(find_emp_code, month, year)
                pay_multi = emp_details.EMP_SAL.DEARANCE_ALLOWANCES * sal.over_time
                Si_No = emp_details.NEW_EMP.ID
                Name = emp_details.NEW_EMP.EMP_NAME
                Rate_of_Wages = emp_details.EMP_SAL.FIXED_SALARY.split("~")[0]
                No_of_Days_worked = sal.day_persent
                Overtime_hours_worked = sal.over_time
                BASIC = emp_details.EMP_SAL.BASIC.split("~")[0]
                Special_Basic = 0
                DA = emp_details.EMP_SAL.DEARANCE_ALLOWANCES.split("~")[0]
                Payments_overtime = pay_multi
                HRA = emp_details.EMP_SAL.HOUSE_RENT_ALLOWANCES.split("~")[0]
                Earn_Others = emp_details.EMP_SAL.OTHER_ALLOWANCES.split("~")[0]

                total_ear = emp_details.EMP_SAL.BASIC.split("~")[0] + \
                            emp_details.EMP_SAL.DEARANCE_ALLOWANCES.split("~")[
                                0] + \
                            emp_details.EMP_SAL.HOUSE_RENT_ALLOWANCES.split("~")[0] + \
                            emp_details.EMP_SAL.OTHER_ALLOWANCES.split("~")[0] + pay_multi

                Earn_Total = total_ear
                PF = emp_details.EMP_SAL.PF.split("~")[0]
                ESIC = emp_details.EMP_SAL.ESIC.split("~")[0]
                Society = 0
                Income_Tax = emp_details.EMP_SAL.INCOME_TAX.split("~")[0]
                Insurance = 0
                Ded_Others = emp_details.EMP_SAL.OTHER_DEDUCTION.split("~")[0]
                Recoveries = 0

                total_ear = emp_details.EMP_SAL.PF.split("~")[0] + emp_details.EMP_SAL.ESIC.split("~")[0] \
                            + emp_details.EMP_SAL.INCOME_TAX.split("~")[0] + \
                            emp_details.EMP_SAL.OTHER_DEDUCTION.split("~")[0]

                Ded_Total = total_ear
                Net_Payment = 0
                Welfare_Fund = 0
                Transaction_Id = 0
                Date_of_Payment = 0
                Remarks = 0

            report_array.append(details_class)

    except EMP_POLICE_VERFICATION.DoesNotExist:
        print("VALUE NOT FOUNT IN DB", year, month)
        report_array.append(details_class)

    return report_array


@login_required(login_url='login')
def Form_A(request):
    if request.method == "POST":
        client = request.POST.get('client')
        unit = request.POST.get('unit')
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=client,
                                                                               EMP_COMPANY__UNIT=unit)
        return render(request, 'Form_A.html', {'data': data, 'error_message': data})

    else:
        country_list = Country.objects.all()
        client = CLIENT_DETAILS.objects.all()
        form = request.POST
        stu = {
            'country_list': country_list,
            'client': client,
        }
    return render(request, 'Form_A.html', stu)


def find_emp_code(emp_id, month, year):
    class montly_pay:
        day_persent = 0
        over_time = 0
        shift_time = 0

    try:
        emp_code = Monthly_Attendance_Table.objects.get(EMP_CODE=emp_id, MONTH=month,
                                                        YEAR=year)
        day_persent = emp_code.DAYS_PRESENT
        over_time = emp_code.OVERTIME_HRS
        shift_time = emp_code.SHIFT_ALLOWANCES_HRS

    except Monthly_Attendance_Table.DoesNotExist:
        print("NO DATA VALUE")

    return montly_pay


def export_movies_to_xlsx(request):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    category_queryset = Country.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('ID', 8),
        ('Title', 40),
        ('Description', 80),
        ('Length', 15),
        ('Rating', 15),
        ('Price', 15),
    ]

    # Iterate through movie categories
    for category_index, category in enumerate(category_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=category.Country_Name,
            index=category.ID,
        )
        # Define the background color of the header cells
        fill = PatternFill(
            start_color="FF5733",
            end_color="FF5733",
            fill_type='solid',
        )
        row_num = 1

        # Assign values, styles, and formatting for each cell in the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies of a category
        for category_index, category in enumerate(category_queryset):
            row_num += 1

            # Define data and formats for each cell in the row
            row = [
                ("movie.pk", 'Normal'),
                ("movie.title", 'Normal'),
                ("movie.description", 'Normal'),
                ("timedelta(minutes=movie.length_in_minutes)", 'Normal'),
                ("movie.rating / 100", 'Percent'),
                ("movie.price", 'Currency'),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                if col_num == 4:
                    cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment

        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # set tab color
        worksheet.sheet_properties.tabColor = "FF5733"

    workbook.save(response)

    return response


def getUser(request):
    if request.method == 'GET' and request.is_ajax():
        # country_name = request.POST['country_name']
        # country_name = request.GET['cnt']
        unit_name = request.GET.get('unit_name', None)
        result_set = []
        selected_client = EMP_COMPANY_DETAILS.objects.filter(UNIT=unit_name)
        for units in selected_client:
            name = str(find_emp_name(units.EMP_CODE))
            employee = name + "------>" + units.EMP_CODE
            result_set.append({'name': employee})

        return HttpResponse(simplejson.dumps(result_set), content_type='application/json')
        # return JsonResponse(result_set,status = 200)

    else:
        return redirect('/')


@login_required(login_url='login')
def Loan_Apply(request):
    client = Client.objects.all()

    if request.method == "POST":
        if request.POST.get("GetData") == "GetData":
            employee_code = request.POST.get('employee_list').split("------>")[1]
            employee_name = request.POST.get('employee_list').split("------>")[0]
            stu = {

                'client': client,
                'loan': "yes",
                'employee_code': employee_code,
                'employee_name': employee_name,
                'client': request.POST.get('client'),
                'unit': request.POST.get('unit'),
            }
            return render(request, 'Loan_apply.html', stu)
        elif request.POST.get("GetData") == "UPDATE":
            today = date.today()

            apply_loan = Loan_Table(
                EMP_CODE=request.POST.get("emp_code"),
                EMP_NAME=request.POST.get("emp_name"),
                LOAN_TYPE=request.POST.get("loan_type"),
                LOAN_AMOUNT=request.POST.get("loan_amount"),
                NO_OF_INSTALLMENT=request.POST.get("no_of_install"),
                CLIENT=request.POST.get("client"),
                UNIT=request.POST.get("unit"),
                APPLY_DATE=today,
            )
            apply_loan.save()
            stu = {
                'client': client,
                'success': "SUCCESS FULLY LOAN APPLIED " + request.POST.get("emp_name"),
            }
            return render(request, 'Loan_apply.html', stu)
    else:
        stu = {

            'client': client,
        }
        return render(request, 'Loan_apply.html', stu)

@login_required(login_url='login')
def Form_C(request):
    if request.method == "POST":
        client = request.POST.get('client')
        unit = request.POST.get('unit')
        client_list = []
        country_list = []
        report_array = []
        filter_value = "no"
        client = CLIENT_DETAILS.objects.all()
        for i in range(len(client)):
            variable = client[i].CLIENT
            print(variable, "variable")
            if variable not in client_list:
                client_list.append(variable)
        filter_client = request.POST.get('client')
        result_set = []
        filter_client_list = CLIENT_DETAILS.objects.filter(CLIENT=filter_client)
        for client_details in filter_client_list:
            result_set.append(client_details)

        for i in range(len(result_set)):
            country = result_set[i].COUNTRY_NAME
            if country not in country_list:
                country_list.append(country)
        form = request.POST
        data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                             'NEW_EMP').filter(EMP_COMPANY__CLIENT=filter_client)
        for details in data:
            print(details.EMP_CODE,"jjjj")
            loan_details = EMP_LOAN_DETAILS.objects.filter(EMP_CODE=details.EMP_CODE)
            for loan in loan_details:
                filter_value = "yes"
                class details_class:
                    EMP_CODE = loan.EMP_CODE
                    EMP_NAME= loan.EMP_NAME
                    DESIGNATION = loan.DESIGNATION
                    LOAN_NAME= loan.LOAN_NAME
                    LOAN_DATE = loan.LOAN_ADDED_DATE

                report_array.append(details_class)

        value = {
            'data': report_array,
            'client': client_list,
            'error_message': data,
            'country_list': country_list,
            'filter_value': filter_value,
            'dummy_value': filter_client,
        }
        return render(request, 'emp_loan_reports.html', value)

    else:
        country_list = Country.objects.all()
        client = Client.objects.all()
        form = request.POST
        stu = {
            'country_list': country_list,
            'client': client,
        }
        if request.GET.get('client') is None:
            newlist = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in newlist:
                    newlist.append(variable)
            print(newlist)
            form = request.POST
            stu = {
                'client': newlist,
            }
        else:
            newlist = []
            country_list = []
            client = CLIENT_DETAILS.objects.all()
            for i in range(len(client)):
                variable = client[i].CLIENT
                print(variable)
                if variable not in newlist:
                    newlist.append(variable)
            filter_value = "yes"
            client = request.GET.get('client')
            result_set = []
            filter_client = CLIENT_DETAILS.objects.filter(CLIENT=client)
            for country in filter_client:
                result_set.append(country.COUNTRY_NAME)

            for i in range(len(result_set)):
                country_list = [result_set[i]]
                for e in result_set:
                    if e not in country_list:
                        country_list.append(e)
            print(newlist)
            data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                 'NEW_EMP').filter(EMP_COMPANY__CLIENT=client)
            stu = {
                'data': data,
                'client': newlist,
                'error_message': data,
                'country_list': country_list,
                'filter_value': filter_value,
                'dummy_value': client,
            }
    return render(request, 'emp_loan_reports.html', stu)
    #return render(request, 'Form_C.html', stu)



@login_required(login_url='login')
def Form_XXIII(request):
    countries = Country.objects.all()
    client = Client.objects.all()
    client_list = []
    country_list = []
    report_array = []
    filter_value = "no"
    client = CLIENT_DETAILS.objects.all()
    for i in range(len(client)):
        variable = client[i].CLIENT
        print(variable, "variable")
        if variable not in client_list:
            client_list.append(variable)
    if request.method == "POST":
        if request.POST.get("GetData") == "GetData":
            client_value = request.POST.get("client")
            unit = request.POST.get("unit")
            try:
                data = EMP_POLICE_VERFICATION.objects.select_related('EMP_BANK', 'EMP_COMM', 'EMP_PER', 'EMP_COMPANY',
                                                                     'EMP_SAL',

                                                                     'NEW_EMP').filter(EMP_COMPANY__CLIENT=client_value)
                em_mon_report = []
                count = 0
                for details in data:
                    count += 1
                    print(count)
                    ext = request.POST.get("date_month").split("-")
                    report = find_emp_monthly_atten(details.EMP_CODE, ext[1], ext[0])
                    if report.OVERTIME_HRS >= 1:
                        class details_class:
                            ID = count
                            EMP_CODE = details.EMP_CODE
                            EMP_NAME = details.NEW_EMP.EMP_NAME
                            FATHER_NAME = details.EMP_PER.FATHER_NAME
                            SEX = details.NEW_EMP.GENDER
                            DESIGNATION = details.EMP_COMPANY.DESIGNATION
                            DATE = request.POST.get("date_month").split("-")
                            OVERTIME = report.OVERTIME_HRS
                            NORMAL_RATE_WAGES = details.EMP_SAL.OVERTIME_AMOUNT

                            EARNINGS = float(report.OVERTIME_HRS )* float(details.EMP_SAL.OVERTIME_AMOUNT)

                        em_mon_report.append(details_class)

                val_len = len(em_mon_report)
                if val_len >= 1:
                    return render(request, 'Form_23.html',
                                  {'countries': countries, 'client': client, 'form': request.POST,
                                   'em_mon_report': em_mon_report})
                else:
                    return render(request, 'Form_23.html',
                                  {'countries': countries, 'client': client_list, 'form': request.POST,
                                   'error': "VALUE NOT FOUND, RESET AND SEARCH AGAIN"})

            except EMP_POLICE_VERFICATION.DoesNotExist:

                return render(request, 'Form_23.html',
                              {'countries': countries, 'client': client, 'form': request.POST,
                               'error': "VALUE NOT FOUND"})

        return render(request, 'Form_23.html',
                      {'countries': countries, 'client': client, 'form': request.POST})

    return render(request, 'Form_23.html',
                  {'countries': countries, 'client': client_list, 'form': request.POST})
